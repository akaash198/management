import csv
import html
import openpyxl
from io import BytesIO
from datetime import date
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from .models import Task
from apps.audit.models import AuditLog

class ProjectExporter:
    def __init__(self, project, requested_by, export_format):
        self.project = project
        self.user = requested_by
        self.format = export_format

    def export(self) -> HttpResponse:
        # Log to Audit
        AuditLog.log(
            actor=self.user, 
            action="export", 
            instance=self.project, 
            changes={"format": self.format}
        )

        if self.format == "csv": return self._export_csv()
        if self.format == "xlsx": return self._export_xlsx()
        if self.format == "pdf": return self._export_pdf()
        return HttpResponse("Invalid format", status=400)

    def _get_tasks_queryset(self):
        return Task.objects.filter(
            project=self.project, is_archived=False
        ).select_related(
            "column", "assignee", "reporter"
        ).prefetch_related(
            "labels", "subtasks", "timelogs"
        ).order_by("column__order", "order")

    HEADERS = ["Title", "Column", "Priority", "Assignee", "Reporter", 
               "Due Date", "Completed", "Labels", "Subtasks", "Hours Logged", "Created At"]

    def _task_to_row(self, task) -> list:
        return [
            task.title,
            task.column.name,
            task.priority,
            task.assignee.full_name if task.assignee else "",
            task.reporter.full_name,
            task.due_date.isoformat() if task.due_date else "",
            "Yes" if task.column.is_done_column else "No",
            ", ".join(l.name for l in task.labels.all()),
            f"{task.subtasks.filter(is_completed=True).count()}/{task.subtasks.count()}",
            sum(t.minutes for t in task.timelogs.all()) // 60,
            task.created_at.isoformat(),
        ]

    def _export_csv(self) -> HttpResponse:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{self.project.name}_tasks.csv"'
        writer = csv.writer(response)
        writer.writerow(self.HEADERS)
        for task in self._get_tasks_queryset():
            writer.writerow(self._task_to_row(task))
        return response

    def _export_xlsx(self) -> HttpResponse:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tasks"
        ws.append(self.HEADERS)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="6366F1")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        priority_colors = {
            "urgent": "FEE2E2", "high": "FEF3C7",
            "normal": "EFF6FF", "low": "F0FDF4"
        }
        for task in self._get_tasks_queryset():
            ws.append(self._task_to_row(task))
            fill_color = priority_colors.get(task.priority, "FFFFFF")
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=fill_color)

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.project.name}_tasks.xlsx"'
        wb.save(response)
        return response

    def _export_pdf(self) -> HttpResponse:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                topMargin=1.5*cm, bottomMargin=1.5*cm,
                                leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles = getSampleStyleSheet()
        
        cell_style = ParagraphStyle(
            'TableCellStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10
        )
        header_cell_style = ParagraphStyle(
            'TableHeaderCellStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=colors.white
        )

        elements = []

        escaped_project_name = html.escape(self.project.name)
        escaped_user_name = html.escape(self.user.full_name)

        elements.append(Paragraph(f"{escaped_project_name} — Task Export", styles["Title"]))
        elements.append(Paragraph(
            f"Generated by {escaped_user_name} on {date.today().isoformat()} "
            f"· {self._get_tasks_queryset().count()} tasks",
            styles["Normal"]
        ))
        elements.append(Spacer(1, 0.5*cm))

        col_widths = [5*cm, 2.5*cm, 2*cm, 3*cm, 3*cm, 2.5*cm]
        # Only take first 6 headers for PDF to avoid overflow
        table_data = []
        
        # Header row
        header_row = [Paragraph(html.escape(h), header_cell_style) for h in self.HEADERS[:6]]
        table_data.append(header_row)
        
        for task in self._get_tasks_queryset():
            row = self._task_to_row(task)[:6]
            table_row = [Paragraph(html.escape(str(cell or "")), cell_style) for cell in row]
            table_data.append(table_row)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#6366F1")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8F9FA")]),
            ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#E5E7EB")),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("PADDING", (0,0), (-1,-1), 6),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{self.project.name}_tasks.pdf"'
        return response
